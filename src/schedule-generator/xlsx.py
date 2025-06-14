import json
import pandas as pd
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# JSON dosyasını yükle
with open("src/final_schedule.json", "r", encoding="utf-8") as f:
    data = json.load(f)

# Slot saatleri
slot_times = {
    "1": "08:30 - 09:15", "2": "09:25 - 10:10", "3": "10:20 - 11:05",
    "4": "11:15 - 12:00", "5": "12:30 - 13:15", "6": "13:25 - 14:10",
    "7": "14:20 - 15:05", "8": "15:15 - 16:00", "9": "16:10 - 16:55",
    "10": "17:05 - 17:50", "11": "18:00 - 18:45"
}

# Bölüm eşleşmeleri
departments = {
    "ElektrikElektronik": "EEE",
    "ElektrikElektronikMuh": "EEE",
    "Makine": "MAC",
    "MakineMuh": "MAC",
    "İnşaat": "CIV",
    "Insaat": "CIV",
    "İnşaatMuh": "CIV",
    "InsaatMuh": "CIV",
    "Endüstri": "IE",
    "Endustri": "IE",
    "EndustriMuh": "IE",
    "EndüstriMuh": "IE",
    "Biyomedikal": "BME",
    "BiyomedikalMuh": "BME"
}

# Günler ve sınıflar
gunler = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
siniflar = ["1. Sınıf", "2. Sınıf", "3. Sınıf", "4. Sınıf"]

# Normalize fonksiyonu
def normalize_text(text):
    return text.replace(" ", "").replace("-", "").replace("ı", "i").replace("İ", "i").lower()

# Tüm dersleri bölümlere göre saklayacak yapı
all_lessons = {dep_code: defaultdict(lambda: defaultdict(dict)) for dep_code in departments.values()}

# Veri işleme
for day, rooms in data.items():
    for room, slots in rooms.items():
        for slot, lesson in slots.items():
            slot_str = str(slot)
            # Slot 0 ise atla
            if slot_str == "0" or slot_str not in slot_times:
                print(f"⚠️ Bilinmeyen slot: {slot_str} - Gün: {day}, Oda: {room}")
                continue

            # Bölüm belirleme:
            if lesson.get("bolum") == "UniElectiveCourse":
                bolum_raw = lesson.get("hocaField", "")
            else:
                bolum_raw = lesson.get("bolum", "")

            bolum_norm = normalize_text(bolum_raw)
            bolum_kodu = None
            for keyword, dep_code in departments.items():
                if normalize_text(keyword) in bolum_norm:
                    bolum_kodu = dep_code
                    break

            if bolum_kodu is None:
                print(f"⚠️ Bölüm bulunamadı: {lesson.get('bolum')} / {lesson.get('hocaField', '')}")
                continue

            # Saat karşılığı
            time_str = slot_times[slot_str]

            # Sınıfı ders kodundan çek (örn: 'EEE 311' → '3' → '3. Sınıf')
            try:
                grade_digit = lesson["dersKodu"].split()[1][0]
                grade = f"{grade_digit}. Sınıf"
            except Exception as e:
                print(f"⚠️ Ders kodundan sınıf bulunamadı: {lesson.get('dersKodu', '')} - {e}")
                continue

            # Ders bilgisi formatı
            ders_bilgisi = (
                f'{lesson["dersKodu"]} {lesson["dersAdi"]} ({room})\n'
                f'{lesson["hocaTitle"]} {lesson["hocaAdi"]} {lesson["hocaSoyadi"].strip()}'
            )

            all_lessons[bolum_kodu][day][time_str][grade] = ders_bilgisi

# Excel dosyalarını üretme ve renklendirme
for dep_code, lessons in all_lessons.items():
    rows = []
    for day in gunler:
        if day in lessons:
            for time in sorted(lessons[day], key=lambda x: int(x.split(":")[0])):  # saatleri sıralı yaz
                row = {"Gün": day, "Saat": time}
                for s in siniflar:
                    row[s] = lessons[day][time].get(s, "")
                rows.append(row)

    df = pd.DataFrame(rows)

    # Excel'e yaz
    filename = f"{dep_code}_Ders_Programi_2025_2026.xlsx"
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Ders Programı")
        ws = writer.sheets["Ders Programı"]

        # Stil ayarları
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Altın sarısı
        class_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Açık sarı
        border_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Beyaz

        for col_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx in range(2, len(df) + 2):
            # Gün ve Saat sütunları ortala
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")

            # Ders bilgisi hücreleri için renkli arka plan ve üstten alt satıra kaydırma (wrap)
            for col_idx in range(3, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell.fill = class_fill
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.fill = border_fill

        # Kolon genişlikleri ayarla
        col_widths = {
            "A": 12,  # Gün
            "B": 13,  # Saat
            "C": 30,  # 1. Sınıf
            "D": 30,  # 2. Sınıf
            "E": 30,  # 3. Sınıf
            "F": 30,  # 4. Sınıf
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

    print(f"✅ {filename} oluşturuldu!")